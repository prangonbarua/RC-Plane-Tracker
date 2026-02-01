"""
Flight Exporter
Exports flight data to Excel with bold peak values
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class FlightExporter:
    def __init__(self, export_dir='flights'):
        self.export_dir = export_dir
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)

    def export_flight(self, flight_id, data_logger):
        """Export a flight to Excel with bold peak values"""
        flight = data_logger.get_flight(flight_id)
        if not flight:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Flight Data"

        # Styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        peak_font = Font(bold=True, color="FF0000")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Flight Summary Section
        ws['A1'] = "FLIGHT SUMMARY"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')

        # Summary data
        summary_data = [
            ("Flight ID:", flight_id),
            ("Start Time:", flight.get('start_time', 'N/A')),
            ("End Time:", flight.get('end_time', 'N/A')),
            ("Duration:", f"{flight.get('duration_seconds', 0)} seconds"),
            ("Peak Speed:", f"{flight.get('peak_speed', 0):.1f} mph"),
            ("Peak Altitude:", f"{flight.get('peak_altitude', 0):.1f} m"),
            ("Total Distance:", f"{flight.get('total_distance', 0):.2f} m"),
        ]

        row = 3
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Add space before data table
        row += 2

        # Data Points Header
        ws[f'A{row}'] = "FLIGHT DATA POINTS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A{row}:I{row}')
        row += 2

        # Column headers
        headers = ['#', 'Time', 'Latitude', 'Longitude', 'Altitude (m)', 'Speed (mph)', 'Satellites', 'RSSI', 'SNR']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        header_row = row
        row += 1

        # Find peak values for highlighting
        data_points = flight.get('data_points', [])
        if data_points:
            peak_speed = max(p['speed'] for p in data_points)
            peak_altitude = max(p['altitude'] for p in data_points)
            peak_speed_row = None
            peak_altitude_row = None

            # Data rows
            for idx, point in enumerate(data_points, 1):
                current_row = row

                # Format timestamp
                timestamp = point.get('timestamp', '')
                if timestamp:
                    try:
                        dt = datetime.fromisoformat(timestamp)
                        timestamp = dt.strftime('%H:%M:%S')
                    except:
                        pass

                row_data = [
                    idx,
                    timestamp,
                    point['latitude'],
                    point['longitude'],
                    point['altitude'],
                    point['speed'],
                    point['satellites'],
                    point.get('rssi', 0),
                    point.get('snr', 0.0)
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

                # Mark peak speed (bold red)
                if point['speed'] == peak_speed and peak_speed_row is None:
                    ws.cell(row=current_row, column=6).font = peak_font
                    peak_speed_row = current_row

                # Mark peak altitude (bold red)
                if point['altitude'] == peak_altitude and peak_altitude_row is None:
                    ws.cell(row=current_row, column=5).font = peak_font
                    peak_altitude_row = current_row

                row += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Save file
        filename = f"flight_{flight_id}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        wb.save(filepath)

        return filepath

    def export_all_flights(self, data_logger):
        """Export all flights to a single Excel file"""
        flights = data_logger.get_all_flights()
        if not flights:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "All Flights"

        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ['Flight ID', 'Date', 'Start Time', 'End Time', 'Duration (s)',
                   'Peak Speed (mph)', 'Peak Altitude (m)', 'Distance (m)']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Find overall peaks
        all_peak_speed = max(f.get('peak_speed', 0) for f in flights) if flights else 0
        all_peak_altitude = max(f.get('peak_altitude', 0) for f in flights) if flights else 0

        peak_font = Font(bold=True, color="FF0000")

        # Data rows
        for row_idx, flight in enumerate(flights, 2):
            start_time = flight.get('start_time', '')
            date_str = ''
            time_str = ''
            if start_time:
                try:
                    dt = datetime.fromisoformat(start_time)
                    date_str = dt.strftime('%Y-%m-%d')
                    time_str = dt.strftime('%H:%M:%S')
                except:
                    pass

            end_time = flight.get('end_time', '')
            end_str = ''
            if end_time:
                try:
                    dt = datetime.fromisoformat(end_time)
                    end_str = dt.strftime('%H:%M:%S')
                except:
                    pass

            row_data = [
                flight['id'],
                date_str,
                time_str,
                end_str,
                flight.get('duration_seconds', 0),
                flight.get('peak_speed', 0),
                flight.get('peak_altitude', 0),
                flight.get('total_distance', 0)
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

                # Bold peak values
                if col == 6 and value == all_peak_speed and all_peak_speed > 0:
                    cell.font = peak_font
                if col == 7 and value == all_peak_altitude and all_peak_altitude > 0:
                    cell.font = peak_font

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Save file
        filename = f"all_flights_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        wb.save(filepath)

        return filepath
